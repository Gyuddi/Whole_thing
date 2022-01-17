from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("연습.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

# for x in range(1,11):
#     for y in range(1,11):
#         print(ws.cell(row=x, column=y).value,end=" ")
#     print()

for x in range(1,ws.max_row+1):
    for y in range(1,ws.max_column+1):
        print(ws.cell(row=x, column=y).value,end=" ")
    print()