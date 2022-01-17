from openpyxl import load_workbook
import random
wb = load_workbook("sample.xlsx")
ws = wb.active

ws.move_range("B1:C11",rows=0,cols=1)
ws["B1"].value = "국어"

for x in range(2,12):
    ws.cell(row = x,column = 2,value =random.randint(1,101) )


wb.save("sample_korean.xlsx")