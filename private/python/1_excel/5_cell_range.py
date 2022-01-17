from openpyxl import Workbook
from random import *
import random

wb = Workbook()
ws = wb.active

ws.append(["번호","영어","수학"])
for i in range(1,11):
    ws.append([i,randint(1,100),randint(1,100)])

col_B = ws["B"]
# print(col_B)
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)
# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         print(cell.coordinate,end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#     print("")

# wb.save("sample.xlsx")

# for row in ws.iter_rows(min_row=2,max_row=11,min_col=2,max_col=3):
#     print(row[0].value,row[1].value)

for column in ws.iter_cols(min_row=2,max_row=11,min_col=2,max_col=3):
    print(column.value)